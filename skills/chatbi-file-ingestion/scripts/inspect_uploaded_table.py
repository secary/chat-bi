from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

SCRIPT_DIR = Path(__file__).resolve().parent
SKILLS_DIR = SCRIPT_DIR.parents[1]
sys.path.insert(0, str(SKILLS_DIR))

from _shared.output import skill_response  # noqa: E402
from table_profile import build_table_profile  # noqa: E402

Schema = Dict[str, Dict[str, str]]


SCHEMAS: Dict[str, Schema] = {
    "sales_order": {
        "order_date": {"type": "date", "label": "订单日期"},
        "region": {"type": "text", "label": "区域"},
        "department": {"type": "text", "label": "部门"},
        "product_category": {"type": "text", "label": "产品类别"},
        "product_name": {"type": "text", "label": "产品名称"},
        "channel": {"type": "text", "label": "渠道"},
        "customer_type": {"type": "text", "label": "客户类型"},
        "sales_amount": {"type": "decimal", "label": "销售额"},
        "order_count": {"type": "int", "label": "订单数"},
        "customer_count": {"type": "int", "label": "客户数"},
        "gross_profit": {"type": "decimal", "label": "毛利"},
        "target_amount": {"type": "decimal", "label": "目标销售额"},
    },
    "customer_profile": {
        "stat_month": {"type": "date", "label": "月份"},
        "region": {"type": "text", "label": "区域"},
        "customer_type": {"type": "text", "label": "客户类型"},
        "new_customers": {"type": "int", "label": "新增客户数"},
        "active_customers": {"type": "int", "label": "活跃客户数"},
        "retained_customers": {"type": "int", "label": "留存客户数"},
        "churned_customers": {"type": "int", "label": "流失客户数"},
    },
}


EXTRA_ALIASES = {
    "日期": "order_date",
    "订单时间": "order_date",
    "统计月份": "stat_month",
    "月度": "stat_month",
    "产品分类": "product_category",
    "产品线": "product_category",
    "成交渠道": "channel",
    "营收": "sales_amount",
    "收入": "sales_amount",
    "利润": "gross_profit",
    "目标": "target_amount",
}


def build_header_map() -> Dict[str, str]:
    mapping = dict(EXTRA_ALIASES)
    for schema in SCHEMAS.values():
        for field, config in schema.items():
            mapping[field] = field
            mapping[config["label"]] = field
    return mapping


HEADER_MAP = build_header_map()


def read_csv(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def read_xlsx(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 Excel 需要安装 openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append({header: value for header, value in zip(headers, row)})
    return headers, data_rows


def read_table(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise ValueError("仅支持 .csv、.xlsx、.xlsm 文件")


def normalize_headers(headers: List[str]) -> Dict[str, str]:
    result = {}
    for header in headers:
        normalized = HEADER_MAP.get(str(header).strip())
        if normalized:
            result[header] = normalized
    return result


def infer_table(fields: List[str]) -> Optional[str]:
    scores = {table: len(set(fields) & set(schema.keys())) for table, schema in SCHEMAS.items()}
    best_table, best_score = max(scores.items(), key=lambda item: item[1])
    return best_table if best_score > 0 else None


def coerce_value(value: Any, value_type: str) -> Tuple[Any, Optional[str]]:
    if value is None or str(value).strip() == "":
        return None, "不能为空"
    text = str(value).strip()
    if value_type == "text":
        return text, None
    if value_type == "int":
        try:
            number = Decimal(text)
            if number != number.to_integral_value():
                return value, "应为整数"
            return int(number), None
        except (InvalidOperation, ValueError):
            return value, "应为整数"
    if value_type == "decimal":
        try:
            return str(Decimal(text)), None
        except InvalidOperation:
            return value, "应为数值"
    if value_type == "date":
        if isinstance(value, datetime):
            return value.date().isoformat(), None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.date().isoformat(), None
            except ValueError:
                pass
        return value, "应为日期"
    return value, None


def validate_rows(
    rows: List[Dict[str, Any]], header_map: Dict[str, str], table: str
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    normalized_rows = []
    errors = []
    schema = SCHEMAS[table]
    reverse_map = {target: source for source, target in header_map.items()}
    for row_index, row in enumerate(rows, start=2):
        normalized = {}
        for field, config in schema.items():
            source_header = reverse_map.get(field)
            if not source_header:
                continue
            value, error = coerce_value(row.get(source_header), config["type"])
            normalized[field] = value
            if error:
                errors.append({"row": row_index, "field": field, "message": error})
        normalized_rows.append(normalized)
    return normalized_rows, errors


def inspect_file(
    path: Path,
    table: Optional[str],
    sample_size: int,
    include_rows: bool,
    question: str = "",
) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{path}")
    headers, rows = read_table(path)
    header_map = normalize_headers(headers)
    normalized_fields = list(header_map.values())
    unknown = [header for header in headers if header not in header_map]
    target_table = table or infer_table(normalized_fields)
    missing: List[str] = []
    normalized_rows: List[Dict[str, Any]] = []
    type_errors: List[Dict[str, Any]] = []
    valid = False
    analysis_mode = "profile_only"
    preview_rows: List[Dict[str, Any]]
    rows_output: List[Dict[str, Any]]
    table_profile = build_table_profile(headers, rows)

    if target_table in SCHEMAS:
        required = set(SCHEMAS[target_table].keys())
        present = set(normalized_fields)
        missing = sorted(required - present)
        normalized_rows, type_errors = validate_rows(rows, header_map, target_table)
        valid = not missing and not type_errors

    if valid and target_table:
        analysis_mode = "schema_validated"
        preview_rows = normalized_rows[:sample_size]
        rows_output = normalized_rows if include_rows else []
        text = (
            f"已读取 {path.name}，匹配到业务表 {target_table}，共 {len(rows)} 行，"
            "字段校验通过，已生成结构画像。"
        )
    else:
        preview_rows = rows[:sample_size]
        rows_output = rows if include_rows else []
        if target_table in SCHEMAS:
            text = (
                f"已读取 {path.name}，候选业务表为 {target_table}，共 {len(rows)} 行，"
                "但字段校验未通过，已生成通用结构画像。"
            )
        else:
            text = (
                f"已读取 {path.name}，共 {len(rows)} 行，未匹配到现有业务表，已生成通用结构画像。"
            )
    # Build column_labels: canonical field name → Chinese display label.
    # For schema-matched tables use the schema label; for generic tables fall back to
    # the original CSV header (which may already be Chinese).
    column_labels: Dict[str, str] = {}
    if target_table in SCHEMAS:
        for field, config in SCHEMAS[target_table].items():
            column_labels[field] = config["label"]
    for original_header, canonical in header_map.items():
        if canonical not in column_labels:
            column_labels[canonical] = original_header

    return skill_response(
        "file_ingestion",
        text,
        data={
            "file": str(path),
            "table": target_table,
            "requested_table": table,
            "row_count": len(rows),
            "headers": headers,
            "normalized_headers": header_map,
            "column_labels": column_labels,
            "missing_columns": missing,
            "unknown_columns": unknown,
            "type_errors": type_errors[:50],
            "preview_rows": preview_rows,
            "rows": rows_output,
            "valid": valid,
            "analysis_mode": analysis_mode,
            "table_profile": table_profile,
            "analysis": {
                "summary_title": "表结构画像",
                "shape": {"rows": len(rows), "columns": len(headers)},
                "columns": headers,
            },
        },
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect an uploaded ChatBI CSV/XLSX file.")
    parser.add_argument("file_path")
    parser.add_argument("--table", choices=sorted(SCHEMAS.keys()))
    parser.add_argument("--sample-size", type=int, default=5)
    parser.add_argument("--include-rows", action="store_true")
    parser.add_argument("--question", default="")
    parser.add_argument("--json", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = inspect_file(
            Path(args.file_path),
            args.table,
            args.sample_size,
            args.include_rows,
            question=str(args.question or ""),
        )
    except Exception as exc:
        if args.json:
            print(json.dumps(skill_response("error", str(exc)), ensure_ascii=False))
            return 1
        print(f"错误：{exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(result["text"])
        data = result["data"]
        print(f"缺失字段：{', '.join(data['missing_columns']) or '无'}")
        print(f"未知字段：{', '.join(data['unknown_columns']) or '无'}")
        print(f"类型错误：{len(data['type_errors'])} 个")
        print(json.dumps(data["preview_rows"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
